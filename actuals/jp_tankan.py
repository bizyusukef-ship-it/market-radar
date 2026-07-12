"""日銀「短観（概要）」ページから、大企業製造業DI（業況判断・最近）の実績値と
前回調査の値を取得する。

https://www.boj.or.jp/statistics/tk/gaiyo/{year}/index.htm に年ごとの一覧があり、
各回の結果はZIP（中身はxlsx）で提供される（2026-07-09時点で実ファイルを確認済み）。

このExcelの「計表1」シートは「１．業況判断」の表で、業種ごとに
[大企業・前回調査の最近, 大企業・前回調査の先行き, 大企業・今回調査の最近, ...]
という並びになっている。「製造業」という行が複数回登場する場合があるため、
（要確認・仮定）シート内で最初に出てくる「製造業」行を業況判断の大企業DIとみなす。
"""
import datetime
import io
import zipfile

from bs4 import BeautifulSoup
import openpyxl

from fetchers.common import fetch_url
from . import history_store

EVENT_KEY = "jp_tankan"
UNIT = ""
SHEET_NAME = "計表1"


def _index_url(year):
    return f"https://www.boj.or.jp/statistics/tk/gaiyo/{year}/index.htm"


def _find_latest_zip_url(as_of_date):
    for year in (as_of_date.year, as_of_date.year - 1):
        resp = fetch_url(_index_url(year))
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            continue
        tbody = table.find("tbody")
        if not tbody:
            continue

        best_date = None
        best_url = None
        for tr in tbody.find_all("tr"):
            tds = tr.find_all("td")
            if len(tds) < 3:
                continue
            link = tds[2].find("a")
            if not link or not link.get("href", "").endswith(".zip"):
                continue
            date_text = tds[0].get_text(" ", strip=True).replace("\xa0", "").replace(" ", "")
            import re
            m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", date_text)
            if not m:
                continue
            y, mo, da = (int(x) for x in m.groups())
            release_date = datetime.date(y, mo, da)
            if release_date > as_of_date:
                continue
            if best_date is None or release_date > best_date:
                best_date = release_date
                href = link.get("href")
                best_url = href if href.startswith("http") else f"https://www.boj.or.jp{href}"

        if best_url:
            return best_url
    return None


def fetch_jp_tankan_actual(as_of_date):
    """戻り値: (actual_str_or_None, prior_str_or_None, unit_str)"""
    zip_url = _find_latest_zip_url(as_of_date)
    if zip_url is None:
        return None, history_store.get_prior(EVENT_KEY, as_of_date.isoformat()), UNIT

    resp = fetch_url(zip_url)
    z = zipfile.ZipFile(io.BytesIO(resp.content))
    xlsx_names = [n for n in z.namelist() if n.endswith(".xlsx")]
    if not xlsx_names:
        return None, history_store.get_prior(EVENT_KEY, as_of_date.isoformat()), UNIT

    wb = openpyxl.load_workbook(io.BytesIO(z.read(xlsx_names[0])), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    actual = prior = None
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and str(row[0]).strip() == "製造業":
            prior = row[2]
            actual = row[4]
            break

    if actual is None:
        return None, history_store.get_prior(EVENT_KEY, as_of_date.isoformat()), UNIT

    actual_str = f"+{actual}" if isinstance(actual, (int, float)) and actual > 0 else str(actual)
    history_store.record(EVENT_KEY, as_of_date.isoformat(), actual_str)

    # historyよりもExcel自体に前回値が載っているのでそちらを優先する
    if prior is not None:
        prior_str = f"+{prior}" if isinstance(prior, (int, float)) and prior > 0 else str(prior)
    else:
        prior_str = history_store.get_prior(EVENT_KEY, as_of_date.isoformat())

    return actual_str, prior_str, UNIT


if __name__ == "__main__":
    import sys

    date_str = sys.argv[1] if len(sys.argv) > 1 else datetime.date.today().isoformat()
    d = datetime.date.fromisoformat(date_str)
    actual, prior, unit = fetch_jp_tankan_actual(d)
    print(f"jp_tankan: actual={actual}{unit} prior={prior}{unit}")
