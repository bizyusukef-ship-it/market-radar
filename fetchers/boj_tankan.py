"""日本銀行「統計データの公表・掲載予定」Excel（tkohyos.xlsx）から短観の公表日を取得する。

このファイルは半年ごと（6月末・12月末）に更新され、先行き12か月分の
全統計の公表予定が1つのExcelにまとまっている（2026-07-06時点で実ファイルを確認済み）。

シート「統計データ」内、B列（統計名）に
「短観（全国企業短期経済観測調査）／概要及び要旨」を含む行の直後の行に、
実際の公表日（datetime）がE列以降に並んでいる。同じ行のD列に公表時刻（time）がある。
"""
import datetime
import io

import openpyxl

from .common import fetch_url, jst_naive_to_iso, make_row
from config.impact_map import EVENT_DEFINITIONS

URL = "https://www.boj.or.jp/statistics/outline/tkohyos.xlsx"
SHEET_NAME = "統計データ"
TARGET_LABEL = "短観（全国企業短期経済観測調査）／概要及び要旨"


def fetch_boj_tankan_events():
    resp = fetch_url(URL)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

    warnings = []
    if SHEET_NAME not in wb.sheetnames:
        return [], [f"BOJ短観: シート '{SHEET_NAME}' が見つかりません。ファイル構成が変わった可能性"]

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    events = []
    label_row_idx = None
    for i, row in enumerate(rows):
        label = row[1] if len(row) > 1 else None
        if label and TARGET_LABEL in str(label):
            label_row_idx = i
            break

    if label_row_idx is None:
        return [], [f"BOJ短観: ラベル '{TARGET_LABEL}' を含む行が見つかりません。ファイル構成が変わった可能性"]

    time_row = rows[label_row_idx]
    date_row = rows[label_row_idx + 1] if label_row_idx + 1 < len(rows) else None
    if date_row is None:
        return [], ["BOJ短観: 日付が入っているはずの次の行がありません"]

    hour, minute = 8, 50
    time_cell = time_row[3] if len(time_row) > 3 else None
    if isinstance(time_cell, datetime.time):
        hour, minute = time_cell.hour, time_cell.minute
    else:
        warnings.append(f"BOJ短観: 公表時刻セルが想定形式(time型)でないため既定値 {hour}:{minute:02d} を使用")

    for cell in date_row:
        if isinstance(cell, datetime.datetime):
            d = cell.date()
            dt_jst = jst_naive_to_iso(d, hour, minute)
            events.append(make_row("jp_tankan", EVENT_DEFINITIONS["jp_tankan"], dt_jst))

    return events, warnings


if __name__ == "__main__":
    evs, warns = fetch_boj_tankan_events()
    for e in evs:
        print(e["datetime_jst"], e["name"])
    for w in warns:
        print("WARNING:", w)
